#!/usr/bin/env python3
import base64
import json
import secrets
import sys
from hashlib import sha256
from html import escape
from pathlib import Path

import numpy

# Compatibility shim for the locally installed openpyxl version.
numpy.float = float
numpy.int = int
numpy.bool = bool

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
DEFAULT_SOURCE_XLSX = PROJECT_ROOT / "dummy_1800.xlsx"
DEFAULT_OUTPUT_HTML = PROJECT_ROOT / "html" / "lookup.html"
PBKDF2_ITERATIONS = 30_000


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>{title}</title>
  <style>
    :root {{
      --bg-1: #f4efe6;
      --bg-2: #dbe7f1;
      --card: rgba(255, 252, 247, 0.92);
      --line: rgba(16, 24, 40, 0.1);
      --text: #132238;
      --muted: #526071;
      --accent: #0c6b58;
      --accent-2: #d16b3d;
      --danger: #b42318;
      --shadow: 0 24px 60px rgba(19, 34, 56, 0.16);
    }}

    * {{
      box-sizing: border-box;
    }}

    body {{
      margin: 0;
      min-height: 100vh;
      font-family: "Pretendard Variable", "Apple SD Gothic Neo", "Noto Sans KR", sans-serif;
      color: var(--text);
      background:
        radial-gradient(circle at top left, rgba(255, 255, 255, 0.8), transparent 35%),
        radial-gradient(circle at bottom right, rgba(209, 107, 61, 0.18), transparent 30%),
        linear-gradient(135deg, var(--bg-1), var(--bg-2));
      padding: 24px;
      display: grid;
      place-items: center;
    }}

    .shell {{
      width: min(100%, 760px);
      position: relative;
    }}

    .shell::before,
    .shell::after {{
      content: "";
      position: absolute;
      inset: auto;
      border-radius: 999px;
      filter: blur(18px);
      z-index: 0;
    }}

    .shell::before {{
      width: 180px;
      height: 180px;
      background: rgba(12, 107, 88, 0.18);
      top: -40px;
      right: 40px;
    }}

    .shell::after {{
      width: 140px;
      height: 140px;
      background: rgba(209, 107, 61, 0.18);
      bottom: -30px;
      left: 20px;
    }}

    .card {{
      position: relative;
      z-index: 1;
      backdrop-filter: blur(14px);
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 28px;
      box-shadow: var(--shadow);
      overflow: hidden;
    }}

    .hero {{
      padding: 32px 32px 20px;
      border-bottom: 1px solid var(--line);
      background:
        linear-gradient(120deg, rgba(12, 107, 88, 0.08), rgba(209, 107, 61, 0.08)),
        rgba(255, 255, 255, 0.55);
    }}

    .eyebrow {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      padding: 8px 12px;
      border-radius: 999px;
      background: rgba(12, 107, 88, 0.08);
      color: var(--accent);
      font-size: 13px;
      font-weight: 700;
      letter-spacing: 0.02em;
    }}

    h1 {{
      margin: 16px 0 10px;
      font-size: clamp(28px, 5vw, 42px);
      line-height: 1.05;
      letter-spacing: -0.04em;
    }}

    .intro {{
      margin: 0;
      color: var(--muted);
      font-size: 16px;
      line-height: 1.7;
    }}

    .content {{
      padding: 28px 32px 32px;
    }}

    form {{
      display: grid;
      gap: 14px;
    }}

    label {{
      font-size: 14px;
      font-weight: 700;
    }}

    .input-row {{
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 12px;
    }}

    input {{
      width: 100%;
      padding: 16px 18px;
      border-radius: 16px;
      border: 1px solid rgba(19, 34, 56, 0.14);
      font-size: 16px;
      color: var(--text);
      background: rgba(255, 255, 255, 0.88);
      outline: none;
      transition: border-color 160ms ease, box-shadow 160ms ease, transform 160ms ease;
    }}

    input:focus {{
      border-color: rgba(12, 107, 88, 0.62);
      box-shadow: 0 0 0 4px rgba(12, 107, 88, 0.12);
      transform: translateY(-1px);
    }}

    button {{
      border: 0;
      padding: 0 22px;
      border-radius: 16px;
      background: linear-gradient(135deg, var(--accent), #0b7e66);
      color: #fff;
      font-size: 15px;
      font-weight: 800;
      cursor: pointer;
      transition: transform 160ms ease, box-shadow 160ms ease, opacity 160ms ease;
      box-shadow: 0 14px 28px rgba(12, 107, 88, 0.22);
    }}

    button:hover {{
      transform: translateY(-1px);
    }}

    button:focus-visible {{
      outline: none;
      box-shadow:
        0 0 0 4px rgba(12, 107, 88, 0.14),
        0 14px 28px rgba(12, 107, 88, 0.22);
    }}

    button:disabled {{
      opacity: 0.7;
      cursor: wait;
    }}

    .hint {{
      margin: 4px 0 0;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.6;
    }}

    .status {{
      min-height: 0;
      margin: 18px 0 0;
      padding: 0;
      font-size: 14px;
      font-weight: 700;
      line-height: 1.6;
      border-radius: 18px;
      transition: all 180ms ease;
    }}

    .status.error {{
      padding: 18px 20px;
      color: #8b3a2b;
      background: linear-gradient(135deg, #fff7f5, #fceae5);
      border: 1px solid rgba(180, 35, 24, 0.14);
      box-shadow: 0 8px 20px rgba(180, 35, 24, 0.08);
      font-size: 17px;
      font-weight: 700;
      letter-spacing: -0.02em;
      text-align: center;
    }}

    .status.ok {{
      color: var(--accent);
    }}

    .result {{
      margin-top: 22px;
      display: none;
      border: 1px solid var(--line);
      border-radius: 22px;
      background: rgba(255, 255, 255, 0.75);
      overflow: hidden;
    }}

    .result.visible {{
      display: block;
      animation: reveal 240ms ease;
    }}

    .result-head {{
      padding: 18px 20px;
      border-bottom: 1px solid var(--line);
      background: rgba(19, 34, 56, 0.04);
    }}

    .result-head strong {{
      display: block;
      font-size: 18px;
    }}

    .result-head span {{
      color: var(--muted);
      font-size: 14px;
    }}

    .result-body {{
      padding: 18px;
      display: grid;
      gap: 16px;
    }}

    .result-meta {{
      display: flex;
      align-items: center;
      gap: 10px;
      padding: 14px 16px;
      border-radius: 16px;
      background: rgba(19, 34, 56, 0.04);
      border: 1px solid rgba(19, 34, 56, 0.08);
    }}

    .result-meta-label {{
      color: var(--muted);
      font-size: 13px;
      font-weight: 700;
    }}

    .result-meta-value {{
      display: inline-flex;
      align-items: center;
      min-height: 34px;
      padding: 6px 12px;
      border-radius: 999px;
      background: rgba(19, 34, 56, 0.08);
      color: var(--text);
      font-size: 14px;
      font-weight: 800;
    }}

    .grid {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      border: 1px solid rgba(19, 34, 56, 0.08);
      border-radius: 18px;
      overflow: hidden;
      background: rgba(255, 255, 255, 0.92);
    }}

    .info-note {{
      margin-top: 18px;
      padding: 18px 20px;
      background: rgba(255, 255, 255, 0.6);
      border: 1px solid rgba(19, 34, 56, 0.08);
      border-radius: 18px;
    }}

    .info-note[hidden] {{
      display: none;
    }}

    .info-note strong {{
      display: block;
      margin-bottom: 10px;
      font-size: 14px;
    }}

    .info-note ul {{
      margin: 0;
      padding-left: 18px;
    }}

    .info-note li {{
      color: var(--muted);
      font-size: 14px;
      line-height: 1.5;
    }}

    .info-note li + li {{
      margin-top: 6px;
    }}

    .cell {{
      padding: 18px;
      border-right: 1px solid var(--line);
    }}

    .cell:last-child {{
      border-right: 0;
    }}

    .cell-label {{
      display: block;
      margin-bottom: 8px;
      color: var(--muted);
      font-size: 13px;
      font-weight: 700;
    }}

    .cell-value {{
      font-size: 18px;
      font-weight: 800;
      word-break: break-word;
    }}

    .cell-value.badge {{
      display: inline-flex;
      align-items: center;
      min-height: 40px;
      padding: 8px 14px;
      border-radius: 999px;
      border: 1px solid transparent;
    }}

    .cell-value.is-registered {{
      color: #0f5132;
      background: #dcfce7;
      border-color: #86efac;
    }}

    .cell-value.is-missing {{
      color: #9f1239;
      background: #ffe4e6;
      border-color: #fda4af;
    }}

    .cell-value.is-pending {{
      color: #92400e;
      background: #ffedd5;
      border-color: #fdba74;
    }}

    .cell-value.is-complete {{
      color: #1d4ed8;
      background: #dbeafe;
      border-color: #93c5fd;
    }}

    @keyframes reveal {{
      from {{
        opacity: 0;
        transform: translateY(6px);
      }}
      to {{
        opacity: 1;
        transform: translateY(0);
      }}
    }}

    @media (max-width: 640px) {{
      body {{
        padding: 16px;
      }}

      .hero,
      .content {{
        padding-left: 20px;
        padding-right: 20px;
      }}

      .input-row,
      .grid {{
        grid-template-columns: 1fr;
      }}

      .cell {{
        min-height: auto;
        border-right: 0;
        border-bottom: 1px solid var(--line);
      }}

      .cell:last-child {{
        border-bottom: 0;
      }}
    }}
  </style>
</head>
<body>
  <main class="shell">
    <section class="card">
      <header class="hero">
        <div class="eyebrow">Knox ID 조회</div>
        <h1>{heading}</h1>
        <p class="intro">{description}</p>
      </header>
      <div class="content">
        <form id="lookup-form">
          <label for="employee-id">아이디</label>
          <div class="input-row">
            <input
              id="employee-id"
              name="employee-id"
              type="text"
              autocomplete="username"
              placeholder="Knox ID를 입력하세요"
              required
            />
            <button id="submit-button" type="submit">조회</button>
          </div>
          <p class="hint">아이디가 일치할 때만 대상자 여부와 신청 상태가 표시됩니다.</p>
        </form>

        <p id="status" class="status" aria-live="polite"></p>

        <section id="result" class="result" aria-live="polite">
          <div class="result-head">
            <strong>조회 결과</strong>
            <span id="result-subtitle"></span>
          </div>
          <div class="result-body">
            <div class="result-meta">
              <span class="result-meta-label">사용자그룹</span>
              <span id="result-group" class="result-meta-value"></span>
            </div>
            <div id="result-grid" class="grid"></div>
          </div>
        </section>
        <section id="info-note" class="info-note" aria-label="안내사항" hidden>
          <strong>※ 안내사항</strong>
          <ul>
            <li>신청 완료 후 계정은 익일 생성되며 방화벽 룰은 익일 오후 적용됩니다.(금요일에 신청 완료된 경우 차주 월요일 적용)</li>
            <li>익일 이후에도 접속이 되지 않을 경우 Q&amp;A 게시판에 신청 차수를 남겨 주세요.</li>
            <li>예) SingleID/방화벽 신청상태가 완료(1차-4/17)인 경우, 4/18 계정 생성 후 오후에 방화벽 룰이 적용됩니다.</li>
          </ul>
        </section>
      </div>
    </section>
  </main>

  <script id="lookup-data" type="application/json">{data_json}</script>
  <script>
    const LOOKUP_CONFIG = {{
      iterations: {iterations},
      records: JSON.parse(document.getElementById("lookup-data").textContent)
    }};

    const form = document.getElementById("lookup-form");
    const input = document.getElementById("employee-id");
    const button = document.getElementById("submit-button");
    const statusNode = document.getElementById("status");
    const resultNode = document.getElementById("result");
    const resultGrid = document.getElementById("result-grid");
    const resultGroup = document.getElementById("result-group");
    const resultSubtitle = document.getElementById("result-subtitle");
    const infoNoteNode = document.getElementById("info-note");

    function normalizeId(value) {{
      return value.trim().toLowerCase();
    }}

    function b64ToBytes(value) {{
      const binary = atob(value);
      const bytes = new Uint8Array(binary.length);
      for (let i = 0; i < binary.length; i += 1) {{
        bytes[i] = binary.charCodeAt(i);
      }}
      return bytes;
    }}

    function bytesToBase64Url(buffer) {{
      const bytes = new Uint8Array(buffer);
      let binary = "";
      for (const byte of bytes) {{
        binary += String.fromCharCode(byte);
      }}
      return btoa(binary).replace(/\\+/g, "-").replace(/\\//g, "_").replace(/=+$/g, "");
    }}

    async function lookupToken(text) {{
      const encoded = new TextEncoder().encode(text);
      const digest = await crypto.subtle.digest("SHA-256", encoded);
      return bytesToBase64Url(digest.slice(0, 16));
    }}

    async function deriveKey(normalizedId, saltBytes) {{
      const baseKey = await crypto.subtle.importKey(
        "raw",
        new TextEncoder().encode(normalizedId),
        "PBKDF2",
        false,
        ["deriveKey"]
      );
      return crypto.subtle.deriveKey(
        {{
          name: "PBKDF2",
          salt: saltBytes,
          iterations: LOOKUP_CONFIG.iterations,
          hash: "SHA-256"
        }},
        baseKey,
        {{
          name: "AES-GCM",
          length: 256
        }},
        false,
        ["decrypt"]
      );
    }}

    function renderResult(payload) {{
      const entries = [
        {{
          label: "PoC 대상자 여부",
          value: "대상",
          tone: "registered"
        }},
        {{
          label: "보안서약서 여부",
          value: payload.g === "GeminiAX" ? "개별 징구" : (payload.a ? "완료" : "미완료"),
          tone: payload.g === "GeminiAX" ? "pending" : (payload.a ? "registered" : "missing")
        }},
        {{
          label: "SingleID/방화벽 신청 상태",
          value: payload.g === "GeminiAX" ? "개별 신청" : (payload.f ? `완료(${{payload.f}})` : "미완료"),
          tone: payload.g === "GeminiAX" ? "pending" : (payload.f ? "registered" : "missing")
        }}
      ];

      resultGrid.replaceChildren();
      for (const entry of entries) {{
        const cell = document.createElement("div");
        cell.className = "cell";

        const labelNode = document.createElement("span");
        labelNode.className = "cell-label";
        labelNode.textContent = entry.label;

        const valueNode = document.createElement("div");
        valueNode.className = "cell-value";
        valueNode.textContent = entry.value === "" ? "-" : String(entry.value);
        if (entry.tone) {{
          valueNode.classList.add("badge", `is-${{entry.tone}}`);
        }}

        cell.append(labelNode, valueNode);
        resultGrid.appendChild(cell);
      }}

      resultGroup.textContent = payload.g || "-";
      resultSubtitle.textContent = "";
      resultNode.classList.add("visible");
      infoNoteNode.hidden = false;
    }}

    function setStatus(message, type) {{
      statusNode.textContent = message;
      statusNode.className = `status${{type ? ` ${{type}}` : ""}}`;
    }}

    function clearResult() {{
      resultGrid.replaceChildren();
      resultGroup.textContent = "";
      resultNode.classList.remove("visible");
      resultSubtitle.textContent = "";
      infoNoteNode.hidden = true;
    }}

    async function findRecordById(rawId) {{
      const normalizedId = normalizeId(rawId);
      if (!normalizedId) {{
        throw new Error("아이디를 입력해 주세요.");
      }}

      const lookupKey = await lookupToken(`lookup:${{normalizedId}}`);
      const record = LOOKUP_CONFIG.records[lookupKey];
      if (!record) {{
        throw new Error("입력하신 정보는 외부 AI 서비스 PoC 대상자 명단에 없습니다.");
      }}

      const key = await deriveKey(normalizedId, b64ToBytes(record[0]));
      try {{
        const plaintextBuffer = await crypto.subtle.decrypt(
          {{
            name: "AES-GCM",
            iv: b64ToBytes(record[1])
          }},
          key,
          b64ToBytes(record[2])
        );
        const text = new TextDecoder().decode(plaintextBuffer);
        return JSON.parse(text);
      }} catch (error) {{
        throw new Error("아이디가 올바르지 않거나 복호화에 실패했습니다.");
      }}
    }}

    form.addEventListener("submit", async (event) => {{
      event.preventDefault();
      clearResult();
      setStatus("조회 중입니다...", "");
      button.disabled = true;

      try {{
        const payload = await findRecordById(input.value);
        renderResult(payload);
        setStatus("", "");
      }} catch (error) {{
        setStatus(error.message, "error");
      }} finally {{
        button.disabled = false;
      }}
    }});
  </script>
</body>
</html>
"""


def normalize_id(value):
    return str(value).strip().lower()


def has_value(value):
    return value not in (None, "")


def b64(data):
    return base64.b64encode(data).decode("ascii")
def load_rows(source_path):
    workbook = load_workbook(source_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
      raise ValueError("엑셀에 데이터가 없습니다.")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if not headers or headers[0] == "":
      raise ValueError("첫 번째 컬럼에 아이디 헤더가 필요합니다.")

    id_header = headers[0]
    header_map = {header: index for index, header in enumerate(headers)}
    required_headers = ["보안서약서", "방화벽 차수", "사용자그룹"]
    missing_headers = [header for header in required_headers if header not in header_map]
    if missing_headers:
        raise ValueError(f"필수 컬럼이 없습니다: {', '.join(missing_headers)}")

    result = []

    for row in rows[1:]:
        if row is None:
            continue

        raw_id = row[0] if len(row) > 0 else None
        if raw_id in (None, ""):
            continue

        normalized = normalize_id(raw_id)
        agreement_value = row[header_map["보안서약서"]] if header_map["보안서약서"] < len(row) else None
        firewall_round = row[header_map["방화벽 차수"]] if header_map["방화벽 차수"] < len(row) else None
        user_group = row[header_map["사용자그룹"]] if header_map["사용자그룹"] < len(row) else None

        agreement_registered = bool(agreement_value)
        firewall_requested = has_value(firewall_round)

        result.append({
            "id_header": id_header,
            "normalized_id": normalized,
            "agreement_registered": agreement_registered,
            "firewall_round": str(firewall_round) if firewall_requested else "",
            "user_group": str(user_group).strip() if has_value(user_group) else "",
        })

    return id_header, result


def build_encrypted_records(rows):
    from cryptography.hazmat.backends import default_backend
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    from cryptography.hazmat.primitives import hashes

    records = {}

    for row in rows:
        normalized_id = row["normalized_id"]
        key_name = base64.urlsafe_b64encode(
            sha256(f"lookup:{normalized_id}".encode("utf-8")).digest()[:16]
        ).decode("ascii").rstrip("=")

        salt = secrets.token_bytes(16)
        iv = secrets.token_bytes(12)
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=salt,
            iterations=PBKDF2_ITERATIONS,
            backend=default_backend(),
        )
        key = kdf.derive(normalized_id.encode("utf-8"))

        payload = json.dumps(
            {
                "a": 1 if row["agreement_registered"] else 0,
                "f": row["firewall_round"],
                "g": row["user_group"],
            },
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode("utf-8")
        ciphertext = AESGCM(key).encrypt(iv, payload, None)

        records[key_name] = [b64(salt), b64(iv), b64(ciphertext)]

    return records


def build_html(encrypted_records):
    return HTML_TEMPLATE.format(
        title=escape("외부 AI 서비스 PoC 대상자 조회"),
        heading=escape("외부 AI 서비스 PoC 대상자 조회"),
        description=escape("본인 Knox ID로 대상자 여부 및 신청 상태를 조회해보세요."),
        iterations=PBKDF2_ITERATIONS,
        data_json=json.dumps(encrypted_records, ensure_ascii=False, separators=(",", ":")),
    )


def main():
    source_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_SOURCE_XLSX
    output_path = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else DEFAULT_OUTPUT_HTML
    id_header, rows = load_rows(source_path)
    encrypted_records = build_encrypted_records(rows)
    html = build_html(encrypted_records)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    print(f"source: {source_path}")
    print(f"generated: {output_path}")
    print(f"records: {len(rows)}")


if __name__ == "__main__":
    main()
